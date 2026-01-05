import os
import re
import openpyxl
from pathlib import Path


class LIN_CANoeConfig:

    def __init__(self):
        self.filePathLDF = ""
        self.filePathExcel = ""
        self.output_root = ""
        self.repo_root = Path(__file__).resolve().parents[2]
        self.find_input_paths()

        self.Signals = {}
        self.Messages = {}
        self.Slaves = []

    
    def find_input_paths(self):
        input_root = self.repo_root / Path("inputs") / Path("lin")
        ldf_files = sorted((input_root / "ldf").glob("*.ldf"))
        excel_files = sorted((input_root / "excel").glob("*.xlsx"))
        if not ldf_files or not excel_files:
            raise FileNotFoundError(
                f"Expected .ldf and .xlsx under {input_root}/ldf and {input_root}/excel"
            )
        self.filePathLDF = ldf_files[0]
        self.filePathExcel = excel_files[0]
        
        self.output_root = self.repo_root / Path("outputs") / Path("lin")
        self.output_root.mkdir(parents=True, exist_ok=True)

        print("\n[LIN] CANoe Configuration Auto-Generation")


    def Extract_Messages_E2E(self):
        self.Extract_Messages_Parameters()
        # Open Excel file
        EXfile = openpyxl.load_workbook(self.filePathExcel)
        for sheet in EXfile.sheetnames:
            for slave in self.Slaves:
                # select a sheet that contain slave signals
                if re.search(slave, sheet):
                    # Extract CheckSumInit Const
                    i = 0
                    # FirstRow is the first sheet row containing all columns title
                    FirstRow = list(EXfile[sheet].iter_rows(values_only=True))[0]
                    for ColTitle in FirstRow:
                        i = i + 1
                        if re.search("Checksum", ColTitle):
                            FrameName = None
                            for row in EXfile[sheet]:
                                if row[1].value is not None:
                                    FrameName = row[1].value
                                if FrameName != 'Frame Name':
                                    for signal in self.Messages[FrameName]['signals']:
                                        if re.search("KSUM", signal) and signal == row[3].value:
                                            self.Messages[FrameName]['CheckSumInit'] = row[i-1].value
                            break
                    break


    # Check if the file exists by checking if the path exists or not
    def Check_file_exists(self):
        if os.path.exists(self.filePathLDF):
            print('\nThe file exists')
        else:
            print('\nThe specified file does NOT exist')


    # Printing All signals with their parameters
    def Print_Signals_Parameters(self):
        # Extract all signals with their parameters
        self.Extract_Signals_Parameters()
        for key, data in self.Signals.items():
            print("Signal : ", key, "\n",
            "  Init   :   ", data['init'], "\n",
            "  Min    :   ", data['min'], "\n",
            "  Max    :   ", data['max'], "\n",
            "  desc   : ", data['desc'], "\n\n", )


    # Printing All Messages with their parameters
    def Print_Messages_Parameters(self):
        # Extract all messages with their parameters
        self.Extract_Messages_Parameters()
        # u should Extract all signals with their parameters first
        for key, data in self.Messages.items():
            print("Message  : ", key, "\n",
            "  Slave        :   ", data['slave'], "\n",
            "  ID           :   ", data['id'], "\n",
            #"  Cyclicity   :   ", data['type'], "\n",
            "  Signals      : ", data['signals'], "\n",
            "  ChSumInitCst :   ", data['CheckSumInit'], "\n\n")


    # Extract all signals with their parameters
    def Extract_Signals_Parameters(self):
        # Strings to search
        SG = re.compile(r'^Signals {')
        Encoding = re.compile(r'^Signal_encoding_types {')
        ENC_Ph_Val = re.compile(r'\s+physical_value,')
        ENC_Log_val = re.compile(r'\s+logical_value,')
        end = re.compile(r'^}')
        # Variables
        temp_SG_loop = False
        temp_Encoding_loop = False
        tempEncSgName = 'none'
        done = False
        # Parse the LDF File
        with open(self.filePathLDF, 'r') as LDFfile:
            for line in LDFfile:
                # Signals area start
                if SG.match(line):
                    temp_SG_loop = True
                # Signals_Encoding area start
                elif Encoding.match(line):
                    temp_Encoding_loop = True
                # the end of area '}'
                elif end.match(line):
                    if temp_SG_loop:
                        temp_SG_loop = False
                    elif temp_Encoding_loop:
                        temp_Encoding_loop = False
                # Extracting parameters from Signals Area
                elif temp_SG_loop:
                    # Signals[tempName]
                    SG_name = re.search(r'(.+?):', line)
                    tempName = SG_name.group(1).strip()
                    # 'init': tempInit
                    SG_Init = re.search(r'\s+\d+\,\s+(\d+)', line)
                    SG_Init_ac = re.search(r'\{(.+?)\}', line)
                    if SG_Init_ac:
                        # tempInit = '{' + SG_Init_ac.group(1).strip() + '}'
                        tempInit = str(0)
                    else:
                        tempInit = str(SG_Init.group(1).strip())
                    # Dictionary Signals{}
                    self.Signals[tempName] = {'min': 0, 'max': 0, 'init': tempInit, 'desc': 'none', 'count': 0}
                # Extracting parameters from Signals_Encoding Area
                elif temp_Encoding_loop:
                    ENC_SG = re.search(r'\s+(.+?)_Encoding', line)
                    if ENC_SG:
                        tempEncSgName = ENC_SG.group(1).strip()
                        done = False
                    elif ENC_Ph_Val and not done:
                        done = True
                        ENC_MinMax = re.search(r'(\d+)\,\s+(\d+),', line)
                        for key, data in self.Signals.items():
                            if key == tempEncSgName:
                                data['min'] = ENC_MinMax.group(1).strip()
                                data['max'] = ENC_MinMax.group(2).strip()
                                data['desc'] = ' '
                    elif ENC_Log_val:
                        int_Val = re.search(r'\s+(\d+),', line)
                        str_desc = re.search(r'\"(.+?)\"', line)
                        if int_Val and str_desc:
                            temp_int_Val = int_Val.group(1).strip()
                            temp_str_desc = str_desc.group(1).strip()
                            self.Signals[tempEncSgName]['desc'] += ' ' + str(temp_int_Val) + ' "' + temp_str_desc + '"'
                            self.Signals[tempEncSgName]['count'] += 1
            # Stop Parsing
            LDFfile.close()


    # Extract all messages with their parameters
    def Extract_Messages_Parameters(self):
        # Strings to search
        SL = re.compile("^  Slaves:")
        FL = re.compile(r'^Frames {')
        end = re.compile(r'^}')
        # Variables
        temp_FL_loop = False
        # Parse the ldf File
        with open(self.filePathLDF, 'r') as LDFfile:
            for line in LDFfile:
                if SL.match(line):
                    temp = line.split(':')[1].split(';')[0].strip()
                    self.Slaves = [s.strip() for s in temp.split(',')]
                # Frames area start
                elif FL.match(line):
                    temp_FL_loop = True
                # the end of Frames area
                elif end.match(line):
                    if temp_FL_loop:
                        temp_FL_loop = False
                # Extracting Message parameters
                elif temp_FL_loop:
                    # self.Messages[tempName] (Frames)
                    temp1 = re.search(r'\s+(.+?):', line)
                    # data['signals']
                    temp2 = re.search(r'\s+(.+?),', line)
                    if temp1:
                        tempName = temp1.group(1).strip()
                        # data['id']
                        temp3 = re.search(r'(\d+),', line)
                        tempId = temp3.group(1).strip()
                        # data['slave']
                        temp4 = re.search(r'\,(.+?),', line)
                        tempSlave = temp4.group(1).strip()
                        # Dictionary self.Messages{}
                        self.Messages[tempName] = {'id': tempId, 'slave': tempSlave, 'signals': [], 'CheckSumInit': 0x00}
                    # add signals
                    elif temp2:
                        tempSg = temp2.group(1).strip()
                        self.Messages[tempName]['signals'].append(tempSg)


    # Create Environment Variables File Declaration && Environment Variables File Value Descriptions
    def Generate_EnVars(self):
        print("\n[LIN] Generating Environment Variables")
        envars_root = self.output_root / Path("envars")
        envars_root.mkdir(parents=True, exist_ok=True)
        EnVarDecPath = envars_root / 'EV_Dec.txt'
        EnVarValDescPath = envars_root / 'EV_ValDesc.txt'
        # Create Environment Variables File Declaration
        self.Create_EnVar_File_Dec(EnVarDecPath)
        # Create Environment Variables File Value Descriptions
        self.Create_EnVar_File_ValDesc(EnVarValDescPath)


    # Create Environment Variables File Declaration from Signals
    def Create_EnVar_File_Dec(self, fileName):
        # Extract all signals with their parameters
        self.Extract_Signals_Parameters()
        # Create the Environment Variables
        with open(fileName, 'w') as EnVarDec:
            for key, data in self.Signals.items():
                if data['max'] == 0: data['max'] = 10
                EnVarLine = 'EV_ ENV_' + key + ' : ' + '0' + ' [' + str(data['min']) + '|' + str(data['max']) + '] "" ' + str(data['init']) + ' 329 DUMMY_NODE_VECTOR0 Vector__XXX; \n'
                EnVarDec.write(EnVarLine)
            # Stop Creating
            EnVarDec.close()
        print("[LIN] Environment Variables declaration file created")


    # Create Environment Variables File Value Descriptions from Signals
    def Create_EnVar_File_ValDesc(self, fileName):
        # Extract all signals with their parameters
        self.Extract_Signals_Parameters()
        # Create the Value Descriptions of the Environment Variables
        with open(fileName, 'w') as EnVarValDesc:
            for key, data in self.Signals.items():
                if data['desc'] != "none" and data['count'] == int(data['max']) + 1:
                    EnVarLine = "VAL_ ENV_" + key + data['desc'] + " ;\n"
                    EnVarValDesc.write(EnVarLine)
            # Stop Creating
            EnVarValDesc.close()
        print("[LIN] Environment Variables value descriptions created")


    # Generate the CAPL code from ...
    def Generate_CAPL(self):
        capl_root = self.output_root / Path("capl")
        capl_root.mkdir(parents=True, exist_ok=True)
        # Extract all messages with their parameters
        self.Extract_Messages_Parameters()
        print("\n[LIN] Generating CAPL code")
        # Create CAPL Code File for each virtual slave
        for slave in self.Slaves:
            fileName = slave + ".can"
            filePath = capl_root / Path(fileName)
            with open(filePath, 'w') as File:
                # Add file name to a table
                line = '///////////////////// Code CAPL for : ' + slave + r' \\\\\\\\\\\\\\\\\\\\\\' + '\n\n'
                File.write(line)
                # Variables {}
                File.write('variables\n')
                File.write('{\n')
                for key, data in self.Messages.items():
                    if slave == data['slave']:
                        line = 'linFrame ' + key + ' _m' + key + ';\n'
                        File.write(line)
                        line = 'long RespError' + key + ' =1;\n'
                        File.write(line)
                        line = 'long RespErrorRet' + key + ' =0;\n'
                        File.write(line)
                        line = 'mstimer _tSetResponseError' + key + ';\n'
                        File.write(line)
                        line = 'dword   _' + key + 'CycleTime' + ';\n\n'
                        File.write(line)
                File.write('\n}\n\n\n')
                # On start {}
                File.write('On start \n')
                File.write('{\n')
                for key, data in self.Messages.items():
                    if slave == data['slave']:
                        line = '_' + key + 'CycleTime = 20;\n'
                        File.write(line)
                        line = 'setTimer(_tSetResponseError' + key + ', _' + key + 'CycleTime);\n'
                        File.write(line)
                        for SG in data['signals']:
                            line = '_m' + key + '.' + SG + ' = getValue(ENV_' + SG + ');\n'
                            File.write(line)
                        File.write('\n')
                File.write('\n}\n\n\n')
                # Three CAPL Functions
                for key, data in self.Messages.items():
                    if slave == data['slave']:
                        # On linFrame ... {}
                        line = 'On linFrame ' + key + '\n{\n'
                        File.write(line)
                        line = '\tif(RespError' + key + ' == 1)' + '\n\t{\n'
                        File.write(line)
                        line = '\t\tRespErrorRet' + key + ' = linSetRespError(0);\n\t}\n'
                        File.write(line)
                        line = '\telse\n\t{\n'
                        File.write(line)
                        line = '\t\tRespErrorRet' + key + ' = linSetRespError(1);\n\t}\n'
                        File.write(line)
                        line = '\toutput(_m' + key + ');\n}\n\n'
                        File.write(line)
                        # On Timer ... {}
                        line = 'On timer _tSetResponseError' + key + '\n{\n'
                        File.write(line)
                        line = '\tif(RespError' + key + ' == 1)' + '\n\t{\n'
                        File.write(line)
                        line = '\t\tRespErrorRet' + key + ' = linSetRespError(0);\n\t}\n'
                        File.write(line)
                        line = '\telse\n\t{\n'
                        File.write(line)
                        line = '\t\tRespErrorRet' + key + ' = linSetRespError(1);\n\t}\n'
                        File.write(line)
                        line = '\tsettimer(_tSetResponseError' + key + ',30);\n}\n\n'
                        File.write(line)
                        # on envVar ... {}
                        for SG in data['signals']:
                            line = 'On envVar ENV_' + SG + '\n{\n'
                            File.write(line)
                            line = '_m' + key + '.' + SG + ' = getValue(this);\n'
                            File.write(line)
                            line = 'output(_m' + key + ');\n}\n\n'
                            File.write(line)
                # Stop Generating
                File.close()
        print("[LIN] CAPL code generated for all virtual CAN nodes")


















